"""Excel (.xlsx) generator with Chinese support."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os
from config import OUTPUT_DIR

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_body(ws, row_count, col_count):
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER


def _auto_width(ws, col_count, min_width=12, max_width=40):
    for col in range(1, col_count + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 4)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len, max_width)


def create_schedule_excel(
    filename: str,
    title: str,
    headers: list[str],
    rows: list[list],
    tips: list[str] | None = None,
) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    col_count = len(headers)
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, col_count)

    for r_idx, row_data in enumerate(rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    _style_body(ws, len(rows) + 1, col_count)
    _auto_width(ws, col_count)

    if tips:
        tip_ws = wb.create_sheet("效率建议")
        tip_ws.cell(row=1, column=1, value="序号")
        tip_ws.cell(row=1, column=2, value="建议内容")
        _style_header(tip_ws, 2)
        for i, tip in enumerate(tips, 1):
            tip_ws.cell(row=i + 1, column=1, value=i)
            tip_ws.cell(row=i + 1, column=2, value=tip)
        _style_body(tip_ws, len(tips) + 1, 2)
        _auto_width(tip_ws, 2)

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    return path


def create_data_excel(
    filename: str,
    original_headers: list[str],
    original_rows: list[list],
    result_headers: list[str],
    result_rows: list[list],
    chart_title: str = "",
    chart_type: str = "bar",
    chart_categories_col: int = 1,
    chart_values_col: int = 2,
) -> str:
    wb = Workbook()

    # Result sheet first (active)
    ws_result = wb.active
    ws_result.title = "处理结果"
    col_count_r = len(result_headers)
    for i, h in enumerate(result_headers, 1):
        ws_result.cell(row=1, column=i, value=h)
    _style_header(ws_result, col_count_r)
    for r_idx, row_data in enumerate(result_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws_result.cell(row=r_idx, column=c_idx, value=val)
    _style_body(ws_result, len(result_rows) + 1, col_count_r)
    _auto_width(ws_result, col_count_r)

    # Chart
    if chart_title and len(result_rows) > 0:
        chart = BarChart()
        chart.title = chart_title
        chart.style = 10
        chart.y_axis.title = result_headers[chart_values_col - 1] if chart_values_col <= len(result_headers) else ""
        data = Reference(
            ws_result,
            min_col=chart_values_col,
            min_row=1,
            max_row=len(result_rows) + 1,
        )
        cats = Reference(
            ws_result,
            min_col=chart_categories_col,
            min_row=2,
            max_row=len(result_rows) + 1,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws_result.add_chart(chart, f"{get_column_letter(col_count_r + 2)}2")

    # Original data sheet
    ws_orig = wb.create_sheet("原始数据")
    col_count_o = len(original_headers)
    for i, h in enumerate(original_headers, 1):
        ws_orig.cell(row=1, column=i, value=h)
    _style_header(ws_orig, col_count_o)
    for r_idx, row_data in enumerate(original_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            ws_orig.cell(row=r_idx, column=c_idx, value=val)
    _style_body(ws_orig, len(original_rows) + 1, col_count_o)
    _auto_width(ws_orig, col_count_o)

    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    return path
