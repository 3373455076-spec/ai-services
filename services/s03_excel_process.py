"""Excel数据处理服务 - 输出Excel (.xlsx)"""

import json
import time
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from ai_client import chat
from formatters.excel import create_data_excel
from openpyxl import load_workbook

SYSTEM_PROMPT = (
    "你是数据分析专家。用户提供了Excel数据的表头和样本数据，以及处理需求。"
    "请以纯JSON格式返回处理后的数据，不要包含markdown代码块，"
    "键为：result_headers(结果列名数组), result_rows(二维数组), "
    "chart_title(图表标题), summary(处理说明)"
)


def run():
    print("\nExcel数据处理")
    print("=" * 40)

    file_path = input("请输入Excel文件路径：").strip().strip('"').strip("'")
    if not os.path.isfile(file_path):
        print(f"文件不存在：{file_path}")
        return

    instruction = input("请输入处理需求（一句话描述）：").strip()

    print("\n正在读取Excel文件...")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    rows_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c is not None else "" for c in row]
        else:
            rows_data.append([str(c) if c is not None else "" for c in row])
            if i >= 100:
                break
    wb.close()

    user_msg = (
        f"表头：{json.dumps(headers, ensure_ascii=False)}\n"
        f"数据（共{len(rows_data)}行）：\n"
        f"{json.dumps(rows_data, ensure_ascii=False)}\n\n"
        f"处理需求：{instruction}"
    )

    print("正在处理数据，请稍候...")
    raw = chat(SYSTEM_PROMPT, user_msg)

    data = json.loads(raw)

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    filename = f"数据处理结果_{timestamp}.xlsx"

    path = create_data_excel(
        filename=filename,
        original_headers=headers,
        original_rows=rows_data,
        result_headers=data.get("result_headers", []),
        result_rows=data.get("result_rows", []),
        chart_title=data.get("chart_title", ""),
    )

    summary = data.get("summary", "")
    print(f"\n处理说明：{summary}")
    print(f"结果已生成：{path}")


if __name__ == "__main__":
    run()
